from kivy.uix.screenmanager import ScreenManager
from kivymd.app import MDApp
from kivymd.uix.dialog import MDDialog
from kivymd.uix.pickers import MDDatePicker, MDTimePicker
from kivymd.uix.screen import MDScreen
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.button import MDRaisedButton, MDIconButton, MDFlatButton
from kivymd.uix.label import MDLabel
from kivy.uix.label import Label
from kivy.uix.gridlayout import GridLayout
from kivy.uix.scrollview import ScrollView
from kivymd.uix.textfield import MDTextField
from kivymd.uix.selectioncontrol import MDCheckbox
from kivy.core.window import Window
from kivymd.toast import toast

# Managing dates
from datetime import timedelta, datetime

# File management
import os
import json

# Managing ics link
import requests
from ics import Calendar

# Excel
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from copy import copy

import re

calendar_data = {}
settings_dict = {}


class TaskDialogContent(MDBoxLayout):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.orientation = "vertical"
        self.spacing = "10dp"
        self.size_hint_y = None
        self.height = "300dp"

        self.text_input = MDTextField(
            hint_text="Task Name",
            mode="rectangle",
            size_hint_y=None,
            height=40
        )
        self.add_widget(self.text_input)

        self.date = datetime.now().date()

        self.date_label = MDLabel(
            text=self.date.strftime("%A, %d %B %Y"),
            halign="center",
            theme_text_color="Primary"
        )
        self.add_widget(self.date_label)

        self.date_button = MDRaisedButton(
            text="Change Date",
            on_release=self.open_date_picker
        )
        self.add_widget(self.date_button)

        self.radio_group = MDBoxLayout(orientation='horizontal', spacing=20, padding=10)

        self.untimed_radio = MDCheckbox(group="task_time", active=True)
        self.untimed_radio.bind(active=self.toggle_time_fields)
        self.radio_group.add_widget(self.untimed_radio)
        self.radio_group.add_widget(MDLabel(text="Untimed / All-Day", theme_text_color="Primary"))

        self.timed_radio = MDCheckbox(group="task_time")
        self.timed_radio.bind(active=self.toggle_time_fields)
        self.radio_group.add_widget(self.timed_radio)
        self.radio_group.add_widget(MDLabel(text="Timed", theme_text_color="Primary"))

        self.add_widget(self.radio_group)

        self.start_time_btn = MDRaisedButton(
            text="Select Start Time",
            on_release=self.open_start_time_picker,
            opacity=0,
            disabled=True
        )
        self.end_time_btn = MDRaisedButton(
            text="Select End Time",
            on_release=self.open_end_time_picker,
            opacity=0,
            disabled=True
        )
        self.add_widget(self.start_time_btn)
        self.add_widget(self.end_time_btn)

    def toggle_time_fields(self, _instance, _val):
        is_timed = self.timed_radio.active
        self.start_time_btn.disabled = not is_timed
        self.end_time_btn.disabled = not is_timed
        self.start_time_btn.opacity = 1 if is_timed else 0
        self.end_time_btn.opacity = 1 if is_timed else 0

    def open_date_picker(self, *_):
        date_picker = MDDatePicker()
        date_picker.bind(on_save=self.on_date_selected)
        date_picker.open()

    def on_date_selected(self, _instance, value, _date_range):
        self.date = value
        self.date_label.text = value.strftime("%A, %d %B %Y")

    def open_start_time_picker(self, *_args):
        picker = MDTimePicker()
        picker.bind(time=self.set_start_time)
        picker.open()

    def open_end_time_picker(self, *_args):
        picker = MDTimePicker()
        picker.bind(time=self.set_end_time)
        picker.open()

    def set_start_time(self, _instance, time_obj):
        self.start_time_btn.text = f"Start: {time_obj.strftime('%H:%M')}"

    def set_end_time(self, _instance, time_obj):
        self.end_time_btn.text = f"End: {time_obj.strftime('%H:%M')}"


class WeekViewScreen(MDScreen):  # Week calendar view class
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.dialog = None 
        self.uni_checkbox = MDCheckbox(active=True)
        self.other_checkbox = MDCheckbox(active=True)
        self.link_input = None
        self.task_content = TaskDialogContent()
        self.week_dates = self.get_current_week_dates()
        self.current_monday = self.get_current_week_dates()[0]
        self.static_layout = MDBoxLayout(orientation="vertical", spacing=10, padding=10)  # stores static content like buttons
        self.calendar_dynamic_container = MDBoxLayout(orientation="vertical")
        self.build_week_first_run()

    def build_week_first_run(self):
        # builds header once avoid rerunning when rendering calendar
        static_layout = self.static_layout

        top_row = MDBoxLayout(spacing=10, padding=5, size_hint_y=None, height=60)

        # Text input for iCal link
        self.link_input = MDTextField(
            hint_text="Timetable Link",
            mode="rectangle",
            size_hint_x=0.1,
            size_hint_y=None,
            font_size=12
        )
        top_row.add_widget(self.link_input)

        save_btn = MDRaisedButton(
            text="Save",
            size_hint=(None, None),
            on_release=self.save_settings
        )
        top_row.add_widget(save_btn)

        gen_btn = MDRaisedButton(
            text="Generate Timetable",
            size_hint=(None, None),
            on_release=self.generate_excel
        )
        top_row.add_widget(gen_btn)

        uni_box = MDBoxLayout(orientation="horizontal", spacing=6, size_hint_x=None, width=140)
        uni_box.add_widget(self.uni_checkbox)
        uni_box.add_widget(MDLabel(text="Timetable", halign="left", font_style="Caption"))
        top_row.add_widget(uni_box)

        other_box = MDBoxLayout(orientation="horizontal", spacing=6, size_hint_x=None, width=100)
        other_box.add_widget(self.other_checkbox)
        other_box.add_widget(MDLabel(text="Other", halign="left", font_style="Caption"))
        top_row.add_widget(other_box)

        add_btn = MDIconButton(
            icon="plus",
            pos_hint={"center_y": 0.5},
            on_release=lambda *args: self.show_add_task_dialog(),
            theme_text_color="Custom",
            text_color=MDApp.get_running_app().theme_cls.primary_color
        )
        top_row.add_widget(add_btn)

        prev_btn = MDIconButton(
            icon="chevron-left",
            on_release=self.previous_week,
            theme_text_color="Custom",
            text_color=MDApp.get_running_app().theme_cls.primary_color
        )
        top_row.add_widget(prev_btn)

        next_btn = MDIconButton(
            icon="chevron-right",
            on_release=self.next_week,
            theme_text_color="Custom",
            text_color=MDApp.get_running_app().theme_cls.primary_color
        )
        top_row.add_widget(next_btn)

        static_layout.add_widget(top_row)
        static_layout.add_widget(self.calendar_dynamic_container)  # add
        self.add_widget(static_layout)

        self.build_week()

    def build_week(self):  # builds cells and tasks and takes care of collisions
        self.calendar_dynamic_container.clear_widgets()
        calendar_dynamic_container = self.calendar_dynamic_container

        # Weekday header
        header = GridLayout(cols=8, size_hint_y=None, height=36)
        header.add_widget(MDLabel(text="Time", halign="center"))
        for day in self.week_dates:
            header.add_widget(MDLabel(text=day.strftime("%a\n%d %b"), markup=True, halign="center"))
        calendar_dynamic_container.add_widget(header)

        # Untimed task row
        untimed_row = GridLayout(cols=8, size_hint_y=None, height=60, spacing=4)
        untimed_row.add_widget(MDLabel(text="Untimed/All-Day", halign="center", theme_text_color="Secondary"))

        for d in self.week_dates:
            month_key, day = d.strftime("%m-%Y"), d.strftime("%d")
            tasks = calendar_data.get(month_key, {}).get(day, [])

            untimed_cell = MDBoxLayout(
                orientation="vertical",
                spacing=2,
                padding=4,
                md_bg_color=(0.92, 0.92, 0.92, 1)
            )

            for task in tasks:
                if not ("start_time" in task and "end_time" in task):
                    untimed_cell.add_widget(Label(text=f"• {task['text']}", font_size=10))
                    untimed_cell.md_bg_color = MDApp.get_running_app().theme_cls.primary_light

            untimed_row.add_widget(untimed_cell)

        calendar_dynamic_container.add_widget(untimed_row)

        # Time grid
        grid = GridLayout(cols=8, spacing=4, size_hint_y=None)
        grid.bind(minimum_height=grid.setter("height"))

        # Build 30 min intervals
        interval_times = []
        time = datetime.strptime("08:00", "%H:%M")
        while time <= datetime.strptime("19:30", "%H:%M"):
            interval_times.append(time.strftime("%H:%M"))
            time += timedelta(minutes=30)

        # Render time slots
        for interval in interval_times:
            show_label = interval.endswith(":00")  # Show only on full hour
            grid.add_widget(MDLabel(text=interval if show_label else "", halign="center", theme_text_color="Secondary"))

            for d in self.week_dates:
                month_key, day = d.strftime("%m-%Y"), d.strftime("%d")
                tasks = calendar_data.get(month_key, {}).get(day, [])

                # Filter tasks active at this interval
                active_tasks = []
                current_time = datetime.strptime(interval, "%H:%M")

                for task in tasks:
                    if "start_time" in task and "end_time" in task:
                        start = datetime.strptime(task["start_time"], "%H:%M")
                        end = datetime.strptime(task["end_time"], "%H:%M")
                        if start <= current_time < end:
                            active_tasks.append(task)

                # Create cell and split vertically if needed
                cell = MDBoxLayout(
                    orientation="horizontal",
                    spacing=1,
                    padding=2,
                    height=30,
                    size_hint_y=None,
                    md_bg_color=(0.95, 0.95, 0.95, 1)
                )

                if active_tasks:
                    for task in active_tasks:
                        task_box = MDBoxLayout(
                            orientation="vertical",
                            size_hint_x=(1 / len(active_tasks)),
                            md_bg_color=MDApp.get_running_app().theme_cls.primary_light,
                            padding=2
                        )
                        task_box.add_widget(Label(
                            text=f"{task['text']}\n{task['start_time']}-{task['end_time']}",
                            font_size=9
                        ))
                        cell.add_widget(task_box)

                grid.add_widget(cell)

        scroll = ScrollView()
        scroll.add_widget(grid)
        calendar_dynamic_container.add_widget(scroll)

    def previous_week(self, *_):
        self.current_monday -= timedelta(days=7)
        self.update_week()

    def next_week(self, *_):
        self.current_monday += timedelta(days=7)
        self.update_week()

    def update_week(self):
        self.week_dates = [
            self.current_monday + timedelta(days=i)
            for i in range(7)
        ]
        self.build_week()

    def generate_excel(self, _):
        wb = Workbook()
        ws = wb.active
        ws.title = "Week Schedule"
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1

        # styling
        font_style = Font(size=12)
        border_style = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        start_hour = 8
        end_hour = 20
        time_slots = [
            datetime.strptime(f"{hour}:00", "%H:%M").strftime("%#I:%M %p")  # 12-hour format 11:00 AM
            for hour in range(start_hour, end_hour + 1)
        ]

        weekdays = ["   ", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
        ws.append(weekdays)

        for time in time_slots:
            ws.append([time])
            current_row = ws.max_row  # max_row gets the last row in sheet
            ws.cell(row=current_row, column=1).alignment = Alignment(vertical="top")

        merged_cells = {}
        day_column_map = {  # used to handle collisions
            "Monday": 1,
            "Tuesday": 2,
            "Wednesday": 3,
            "Thursday": 4,
            "Friday": 5
        }
        for d in self.week_dates:  # handle each day in the week
            timetable_clash = False
            weekday = d.strftime("%A")
            if weekday not in weekdays:
                continue  # Skip weekends or unexpected dates

            day_column = day_column_map[weekday]
            month_key, day = d.strftime("%m-%Y"), d.strftime("%d")
            tasks = calendar_data.get(month_key, {}).get(day, [])

            day_tasks_collision_handler = {}
            clashed_tasks = set()  # using set to auto handle duplicate entries

            for task in tasks:  # handle tasks for each day
                exit_var = False
                if "start_time" not in task or "end_time" not in task:
                    continue

                start_time = datetime.strptime(task["start_time"], "%H:%M")
                end_time = datetime.strptime(task["end_time"], "%H:%M")
                task_name = task.get("text", "Untitled")
                task_location = task.get("location")

                current_time = start_time
                while current_time < end_time:
                    task_hour = current_time.hour
                    if task_hour not in day_tasks_collision_handler:
                        day_tasks_collision_handler[task_hour] = [task_name]
                        current_time += timedelta(hours=1)
                    else:
                        day_tasks_collision_handler[task_hour].append(task_name)
                        current_time += timedelta(hours=1)
                        for clashing_task in day_tasks_collision_handler[task_hour]:
                            clashed_tasks.add(clashing_task)

                        if timetable_clash:  # collision already handled, just add times into dictionary
                            continue

                        # collision not handled yet
                        ws.insert_cols(day_column + 2)  # index + 1, then add 1 more to insert column on right

                        # update the day column map
                        for day in day_column_map:
                            if day_column_map[day] >= day_column + 1:
                                day_column_map[day] += 1

                        day_column = day_column_map[weekday] + 1
                        timetable_clash = True

                duration_minutes = (end_time - start_time).seconds // 60
                row_span = duration_minutes // 60

                start_slot_str = start_time.strftime("%#I:%M %p")
                if start_slot_str not in time_slots:
                    continue  # Skip if time doesn't match a slot

                start_row = time_slots.index(start_slot_str) + 2  # +2 for header rows
                end_row = start_row + row_span - 1
                current_cells = [start_row, end_row]

                col_letter = get_column_letter(day_column + 1)  # +1 as Excel col indexes start at 1, eg A = 1

                if col_letter in merged_cells.keys():
                    for cell in current_cells:
                        if cell in merged_cells[col_letter]:
                            exit_var = True
                            break  # If cell has been merged, don't change cell value
                    if exit_var:
                        continue

                # styles used for rich text
                normal_font = InlineFont(sz=12)
                bold_font = InlineFont(b=True, sz=12)

                ws[f"{col_letter}{start_row}"].value = CellRichText(
                                        TextBlock(bold_font, task_name),
                                        TextBlock(normal_font, f"\n{task_location}")
                                    )

                task_type = task_name.split()[-1].lower()
                if task_type in ["lecture", "seminar"]:
                    fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")  # light green
                else:
                    fill = PatternFill(start_color="E6B8B7", end_color="E6B8B7", fill_type="solid")  # light red

                ws[f"{col_letter}{start_row}"].fill = fill

                ws.merge_cells(f"{col_letter}{start_row}:{col_letter}{end_row}")
                if col_letter not in merged_cells:
                    merged_cells[col_letter] = current_cells
                else:
                    for cell in current_cells:
                        merged_cells[col_letter].append(cell)
                ws[f'{col_letter}{start_row}'].alignment = Alignment(
                    vertical='top',
                    wrap_text=True
                )

            # to manage timetable clashes, create another column for clashing tasks and merge all other cells in columns
            if timetable_clash:
                no_merge_hours = set()  # used to store hours when clashes occur

                for hour, tasks in day_tasks_collision_handler.items():
                    for task_name in tasks:
                        if task_name in clashed_tasks:
                            no_merge_hours.add(hour)

                for row in range(1, ws.max_row + 1):
                    if row + 6 in no_merge_hours:
                        continue  # clash so don't merge. row + 6 used to indicate time indices in Excel (starts at 8)

                    start_col_num = day_column_map[weekday] + 1
                    end_col_num = day_column + 1
                    start_col = get_column_letter(start_col_num)
                    end_col = get_column_letter(end_col_num)

                    if self.is_cell_merged(ws, row, start_col_num) or self.is_cell_merged(ws, row, end_col_num):
                        # handle cases where horizontal merge is needed but vertical merge exists
                        for merged_range in list(ws.merged_cells.ranges):
                            if (
                                    merged_range.min_col == merged_range.max_col
                                    and
                                    merged_range.min_row == row
                            ):
                                if start_col_num <= merged_range.min_col <= end_col_num:
                                    # found the entire vertical merge range affecting this row
                                    min_row = merged_range.min_row
                                    max_row = merged_range.max_row
                                    # if top left cell in new range empty, first move data and styling there
                                    if ws[f'{start_col}{min_row}'].value is None or ws[f'{start_col}{min_row}'].value == '':
                                        target = ws[f'{start_col}{min_row}']
                                        source = ws[f'{end_col}{min_row}']

                                        target.value = source.value
                                        target.alignment = copy(source.alignment)
                                        target.fill = copy(source.fill)
                                        target.font = copy(source.font)
                                        target.border = copy(source.border)

                                    # unmerge existing vertical merge first to avoid Excel file corruption
                                    ws.unmerge_cells(str(merged_range))
                                    ws.merge_cells(f"{start_col}{min_row}:{end_col}{max_row}")
                                    break
                        continue

                    ws.merge_cells(f"{start_col}{row}:{end_col}{row}")

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if not isinstance(cell.value, CellRichText):  # don't overwrite cells with bold text
                    cell.font = font_style
                cell.border = border_style

        ws.column_dimensions['A'].width = 10

        for i in range(66, 70 + 1):  # ASCII values
            ws.column_dimensions[chr(i)].width = 20

        for row in range(2, ws.max_row + 1):  # skips header column

            max_height = 15  # default Excel height

            for col in range(2, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)

                if cell.value:
                    text_length = len(str(cell.value))

                    chars_per_line = 22  # rough guess
                    lines = (text_length // chars_per_line) + 1

                    estimated_height = lines * 18  # 18 = height per line

                    if estimated_height > max_height:
                        max_height = estimated_height

            if max_height > 15:
                ws.row_dimensions[row].height = max_height

        # Save file
        try:
            wb.save('Timetable.xlsx')
            MDApp.get_running_app().show_message('Excel timetable generated')
        except Exception as e:
            MDApp.get_running_app().show_message(str(e))

    def is_cell_merged(self, ws, row, col):
        for merged_range in ws.merged_cells.ranges:
            if (
                    merged_range.min_row <= row <= merged_range.max_row
                    and
                    merged_range.min_col <= col <= merged_range.max_col
            ):
                return True
        return False

    def get_current_week_dates(self):
        today = datetime.now()
        monday_current_week = today - timedelta(days=today.weekday())
        return [monday_current_week + timedelta(days=i) for i in range(7)]

    def show_add_task_dialog(self):

        self.dialog = MDDialog(
            title="Add Task",
            text="",
            type="custom",
            content_cls=self.task_content,
            buttons=[
                MDRaisedButton(
                    text="Save",
                    on_release=self.save_task
                ),
                MDFlatButton(
                    text="Cancel",
                    on_release=self.dismiss_dialog
                )
            ]
        )
        self.dialog.open()

    def dismiss_dialog(self, *_args):
        self.dialog.dismiss()

    def save_settings(self, _):
        global settings_dict
        link = self.link_input.text.strip()
        settings_dict['ics_url'] = link
        save_path = os.path.join(MDApp.get_running_app().user_data_dir, "settings.json")  # Compatible file path for all platforms
        try:
            with open(save_path, "w") as file:
                json.dump(settings_dict, file)
        except Exception as e:
            print(f"Failed to save link: {e}")  # Used in event of any errors, then they will be printed

    def save_task(self, *_args):
        global calendar_data

        task_name = self.task_content.text_input.text.strip()
        if not task_name:
            print("Task name is required.")
            return

        date_str_day = self.task_content.date.strftime("%d")  # e.g. "31"
        date_str_month = self.task_content.date.strftime("%m-%Y")  # e.g. "07-2025"

        # Initialise nested dicts if necessary
        if date_str_month not in calendar_data:
            calendar_data[date_str_month] = {}
        if date_str_day not in calendar_data[date_str_month]:
            calendar_data[date_str_month][date_str_day] = []

        # Build task dictionary
        task = {
            "text": task_name,
            "type": "other"
        }

        if self.task_content.timed_radio.active:
            try:
                task["start_time"] = self.task_content.start_time_btn.text.split("Start: ")[1]
                task["end_time"] = self.task_content.end_time_btn.text.split("End: ")[1]
            except IndexError:
                print("Missing time selection")
                return

        # Add task to calendar
        calendar_data[date_str_month][date_str_day].append(task)
        self.dialog.dismiss()
        MDApp.get_running_app().save_data()


class WeekTimetableApp(MDApp):  # Class defines the main app
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.load_data()
        self.week_screen = WeekViewScreen(name="week")

    def build(self):
        self.title = "Week Timetable"
        self.theme_cls.primary_palette = "Blue"
        self.theme_cls.theme_style = "Light"
        Window.clearcolor = (0.98, 0.98, 0.98, 1)

        screen_manager = ScreenManager()
        screen_manager.add_widget(self.week_screen)

        return screen_manager

    def save_data(self):  # Function saves the entered data to a json file in Users/AppData/Roaming/calendar
        save_path = os.path.join(self.user_data_dir, "saved_state.json")  # Compatible file path for all platforms
        try:
            os.makedirs(self.user_data_dir, exist_ok=True)
            with open(save_path, "w") as file:
                json.dump(calendar_data, file, indent=2)  # Indent used for readability of json
        except Exception as e:
            print(f"Failed to save data: {e}")  # Used in event of any errors, then they will be printed

    def load_data(self):  # Function runs to load all data saved in json
        global calendar_data  # Using global calendar data
        try:
            save_path = os.path.join(self.user_data_dir, "saved_state.json")
            if not os.path.exists(save_path):
                return  # If first time launch, json does not exist, so end execution

            with open(save_path, "r") as file:
                calendar_data = json.load(file)  # Retrieve data from json

            self.import_ics_to_calendar_data()  # Load in uni timetable data

        except Exception as e:  # In case of an error
            print(f"Error loading saved data: {e}")

    def import_ics_to_calendar_data(self):
        global settings_dict
        global calendar_data
        settings_path = os.path.join(self.user_data_dir, "settings.json")  # Load link
        if not os.path.exists(settings_path):
            print("No link saved.")
            return False

        try:
            with open(settings_path) as file:
                settings_dict = json.load(file)

            # Fetch and parse
            calendar_request = Calendar(requests.get(settings_dict['ics_url']).text)

            if len(calendar_request.events) > 0:  # Remove all previously loaded uni tasks if link works
                for month_key in list(calendar_data.keys()):
                    for day_key in list(calendar_data[month_key].keys()):
                        calendar_data[month_key][day_key] = [
                            task for task in calendar_data[month_key][day_key]
                            if task.get("type") != "uni"
                        ]

            for event in calendar_request.events:
                request_date = event.begin.strftime("%d")
                month_key = event.begin.strftime("%m-%Y")
                start = event.begin.strftime("%H:%M")
                end = event.end.strftime("%H:%M")

                entry = {
                    "text": event.name,
                    "start_time": start,
                    "end_time": end,
                    "location": self.clean_location(event.location),
                    "type": "uni"
                }

                calendar_data.setdefault(month_key, {}).setdefault(request_date, []).append(entry)
            return True
        except Exception as e:
            print(f"Failed to import: {e}")
            return False

    def clean_location(self, location):
        if not location:
            return "No location"

        if location.strip() == "-":
            return "(Online)"

        parts = location.split(".")
        cleaned_location = f"{parts[0]} {parts[-1]}"

        cleaned_location = re.sub(r"\([^)]*\)", "", cleaned_location)  # remove date ranges in brackets

        return cleaned_location.strip().rstrip(",")

    @staticmethod
    def show_message(message):  # A message is passed which is then displayed by kivy on screen
        toast(message)


WeekTimetableApp().run()
